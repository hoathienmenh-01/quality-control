"""Generate sample Excel report"""
import argparse
from datetime import datetime, timedelta
import random


def generate_report(date_from=None, date_to=None, output="exports/report.xlsx"):
    """Tạo báo cáo Excel mẫu"""
    try:
        import openpyxl
    except ImportError:
        print("❌ Cần cài openpyxl: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # Sheet 1: Tổng quan
    ws1 = wb.active
    ws1.title = "Tổng Quan"
    ws1.append(["BÁO CÁO KIỂM TRA CHẤT LƯỢNG"])
    ws1.append([f"Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws1.append([])
    ws1.append(["Chỉ số", "Giá trị"])
    ws1.append(["Tổng sản phẩm", 1250])
    ws1.append(["PASS", 1180])
    ws1.append(["FAIL", 70])
    ws1.append(["Tỷ lệ đạt", "94.4%"])

    # Sheet 2: Chi tiết lỗi
    ws2 = wb.create_sheet("Chi Tiết Lỗi")
    ws2.append(["Loại lỗi", "Số lượng", "Tỷ lệ"])
    ws2.append(["Thiếu linh kiện", 25, "35.7%"])
    ws2.append(["Sai QR", 18, "25.7%"])
    ws2.append(["Sai SN", 15, "21.4%"])
    ws2.append(["Lệch anten", 12, "17.1%"])

    # Sheet 3: Theo trạm
    ws3 = wb.create_sheet("Theo Trạm")
    ws3.append(["Trạm", "Tổng", "PASS", "FAIL", "Tỷ lệ"])
    ws3.append(["STATION-01", 450, 430, 20, "95.6%"])
    ws3.append(["STATION-02", 400, 375, 25, "93.8%"])
    ws3.append(["STATION-03", 400, 375, 25, "93.8%"])

    import os
    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    print(f"✅ Đã tạo báo cáo: {output}")


def main():
    parser = argparse.ArgumentParser(description="Generate sample Excel report")
    parser.add_argument("--output", default="exports/report.xlsx", help="Output file path")
    args = parser.parse_args()
    generate_report(output=args.output)


if __name__ == "__main__":
    main()
